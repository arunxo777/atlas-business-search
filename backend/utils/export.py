"""JSON, CSV, and Excel export utilities."""

from __future__ import annotations

import csv
import io
import json
from typing import Literal

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import BusinessRecord

ExportFormat = Literal["json", "csv", "xlsx"]


def _business_to_flat(b: BusinessRecord) -> dict[str, str | float | int | None]:
    return {
        "id": b.id,
        "job_id": b.job_id,
        "business_name": b.business_name,
        "address": b.address,
        "phone": "; ".join(b.phone),
        "email": "; ".join(b.email),
        "website": b.website,
        "rating": b.rating,
        "review_count": b.review_count,
        "services": "; ".join(b.services),
        "specialties": "; ".join(b.specialties),
        "license_information": b.license_information,
        "certifications": "; ".join(b.certifications),
        "awards": "; ".join(b.awards),
        "working_hours": (
            "; ".join(f"{k}: {v}" for k, v in b.working_hours.items())
            if b.working_hours
            else None
        ),
        "social_profiles": "; ".join(
            f"{k}={v}" for k, v in b.social_profiles.items()
        ),
        "image_urls": "; ".join(b.image_urls),
        "rank_score": b.rank_score,
        "verification_status": b.verification_status,
        "source_reliability_score": b.source_reliability_score,
        "discovered_at": b.discovered_at.isoformat(),
        "last_updated": b.last_updated.isoformat(),
    }


def export_businesses(
    businesses: list[BusinessRecord],
    fmt: ExportFormat,
) -> tuple[bytes, str, str]:
    """Export businesses to the requested format. Returns (content, media_type, filename)."""
    if fmt == "json":
        data = [b.model_dump(mode="json") for b in businesses]
        content = json.dumps(data, indent=2, default=str).encode("utf-8")
        return content, "application/json", "businesses.json"

    rows = [_business_to_flat(b) for b in businesses]
    if not rows:
        rows = [{"business_name": ""}]

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
        return output.getvalue().encode("utf-8"), "text/csv", "businesses.csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "Businesses"
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return (
        buffer.read(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "businesses.xlsx",
    )
